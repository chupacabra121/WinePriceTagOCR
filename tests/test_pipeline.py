"""Offline tests for geometry, schema, store attribution, dedup and output."""

from pathlib import Path
from types import SimpleNamespace

import pytest

from wine_ocr.images import (
    MAX_EDGE_HIRES, Region, band_regions, estimate_image_tokens, grid_regions,
)
from wine_ocr.models import BandExtraction, PhotoLayout
from wine_ocr.output import COLUMNS, deduplicate, write_csv, write_xlsx
from wine_ocr.schema import json_schema_for
from wine_ocr.stores import load_store_map, resolve_store

# Dimensions of the Annabella sample photos.
PHOTO_W, PHOTO_H = 4284, 5712


def _band(**kw):
    base = dict(
        label="shelf", contains_wine=True, bottles_y0=0.10, bottles_y1=0.26,
        tags_y0=0.28, tags_y1=0.32, x0=0.18, x1=0.80,
    )
    base.update(kw)
    return SimpleNamespace(**base)


# --------------------------------------------------------------------------
# Geometry
# --------------------------------------------------------------------------


def test_tiles_are_never_downscaled():
    """The whole point of tiling: every tile reaches the model at native size."""
    for regions in (
        grid_regions(PHOTO_W, PHOTO_H, MAX_EDGE_HIRES),
        band_regions(PHOTO_W, PHOTO_H, [_band()], MAX_EDGE_HIRES),
    ):
        assert regions
        for r in regions:
            assert max(r.width, r.height) <= MAX_EDGE_HIRES, r


def test_whole_photo_would_be_downscaled():
    """Guards the premise: a single call per photo loses more than half the pixels."""
    assert max(PHOTO_W, PHOTO_H) / MAX_EDGE_HIRES > 2.0


def test_band_spans_bottles_and_their_tag_rail():
    (region,) = band_regions(PHOTO_W, PHOTO_H, [_band(x0=0.3, x1=0.5)], MAX_EDGE_HIRES)
    assert region.y0 < 0.10 * PHOTO_H  # padded above the bottles
    assert region.y1 > 0.32 * PHOTO_H  # and below the tags


def test_non_wine_bands_are_skipped():
    assert band_regions(PHOTO_W, PHOTO_H, [_band(contains_wine=False)]) == []


def test_tiles_overlap_so_nothing_falls_between_them():
    regions = band_regions(PHOTO_W, PHOTO_H, [_band(x0=0.0, x1=1.0)], MAX_EDGE_HIRES)
    assert len(regions) > 1
    for left, right in zip(regions, regions[1:]):
        assert right.x0 < left.x1, "adjacent tiles must overlap"
    assert regions[0].x0 == 0
    assert regions[-1].x1 >= PHOTO_W - 1


def test_degenerate_band_is_dropped():
    assert band_regions(PHOTO_W, PHOTO_H, [_band(x0=0.5, x1=0.5001)]) == []


def test_bbox_maps_from_tile_back_to_whole_photo():
    region = Region("t", 1000, 2000, 2000, 3000, 0, 0, 1)
    x0, y0, x1, y1 = region.to_full(0.0, 0.0, 1.0, 1.0, PHOTO_W, PHOTO_H)
    assert x0 == pytest.approx(1000 / PHOTO_W)
    assert y1 == pytest.approx(3000 / PHOTO_H)
    # A centred box stays centred within the tile's extent.
    cx0, _, cx1, _ = region.to_full(0.25, 0.25, 0.75, 0.75, PHOTO_W, PHOTO_H)
    assert cx0 == pytest.approx(1250 / PHOTO_W)
    assert cx1 == pytest.approx(1750 / PHOTO_W)


def test_token_estimate_respects_tier_ceiling():
    assert estimate_image_tokens(PHOTO_W, PHOTO_H, MAX_EDGE_HIRES) == 4784
    assert estimate_image_tokens(200, 200, MAX_EDGE_HIRES) == pytest.approx(53, abs=2)


# --------------------------------------------------------------------------
# Schema
# --------------------------------------------------------------------------


@pytest.mark.parametrize("model", [PhotoLayout, BandExtraction])
def test_schema_is_compiler_ready(model):
    schema = json_schema_for(model)
    seen_objects = 0

    def walk(node):
        nonlocal seen_objects
        if isinstance(node, dict):
            for banned in ("default", "title", "minimum", "maxLength", "$schema"):
                assert banned not in node, f"{banned} survived cleaning"
            if node.get("type") == "object" and "properties" in node:
                seen_objects += 1
                assert node["additionalProperties"] is False
                assert set(node["required"]) == set(node["properties"])
            for value in node.values():
                walk(value)
        elif isinstance(node, list):
            for value in node:
                walk(value)

    walk(schema)
    assert seen_objects >= 2


def test_optional_fields_are_nullable_not_absent():
    props = json_schema_for(BandExtraction)["$defs"]["WineEntry"]["properties"]
    types = [t.get("type") for t in props["producer"]["anyOf"]]
    assert "null" in types and "string" in types


# --------------------------------------------------------------------------
# Store attribution
# --------------------------------------------------------------------------


@pytest.fixture
def store_map(tmp_path):
    cfg = tmp_path / "stores.yaml"
    cfg.write_text("stores:\n  Annabella:\n    - annabella\n    - anabela\n")
    return load_store_map(cfg)


def test_cli_override_wins(store_map):
    got = resolve_store(
        Path("/p/kaufland/a.jpg"), Path("/p"), "Lidl", store_map, "Carrefour"
    )
    assert (got.store, got.source) == ("Lidl", "cli")
    assert got.read_from_photo == "Carrefour"  # losing candidate is retained


def test_alias_maps_to_canonical_name(store_map):
    got = resolve_store(Path("/p/anabela/a.jpg"), Path("/p"), None, store_map, None)
    assert (got.store, got.source) == ("Annabella", "folder-map")


def test_unmapped_folder_is_prettified(store_map):
    got = resolve_store(Path("/p/mega_image/a.jpg"), Path("/p"), None, store_map, None)
    assert (got.store, got.source) == ("Mega Image", "folder")


def test_generic_and_dateish_folders_fall_through(store_map):
    got = resolve_store(
        Path("/p/photos/2026-05/a.jpg"), Path("/p"), None, store_map, "Profi"
    )
    assert (got.store, got.source) == ("Profi", "photo")


def test_unknown_when_nothing_identifies_the_store(store_map):
    got = resolve_store(Path("/p/photos/a.jpg"), Path("/p"), None, store_map, None)
    assert (got.store, got.source) == ("Unknown", "none")


def test_missing_config_is_not_fatal():
    assert load_store_map(Path("/nope/stores.yaml")) == {}


# --------------------------------------------------------------------------
# Dedup and output
# --------------------------------------------------------------------------


def _row(**kw):
    base = {
        "store": "Annabella", "wine_name": "Cotnari Euforia Busuioaca Roze",
        "vintage": "2023", "volume_ml": 750, "price": 54.59,
        "pairing_confidence": "high", "raw_tag_text": "COTNARI EUFORIA",
        "needs_review": "", "review_reasons": "", "duplicate_of": "",
    }
    base.update(kw)
    return base


def test_duplicates_are_linked_not_deleted():
    rows = deduplicate([_row(), _row(pairing_confidence="low", price=None)])
    assert len(rows) == 2, "no row is ever dropped"
    primary = [r for r in rows if not r["duplicate_of"]]
    assert len(primary) == 1
    # The richer sighting is kept as primary.
    assert primary[0]["price"] == 54.59
    assert rows[1]["duplicate_of"] == primary[0]["row_id"]


def test_conflicting_prices_across_sightings_are_flagged():
    rows = deduplicate([_row(price=54.59), _row(price=58.99)])
    primary = next(r for r in rows if not r["duplicate_of"])
    assert primary["needs_review"] == "yes"
    assert "price disagrees" in primary["review_reasons"]


def test_distinct_vintages_are_not_merged():
    rows = deduplicate([_row(vintage="2022"), _row(vintage="2023")])
    assert len([r for r in rows if not r["duplicate_of"]]) == 2


def test_row_ids_are_unique():
    rows = deduplicate([_row(wine_name=f"Wine {i}") for i in range(5)] + [_row()])
    assert len({r["row_id"] for r in rows}) == len(rows)


def test_writers_produce_readable_files(tmp_path):
    import csv

    rows = deduplicate([_row(), _row(wine_name="Purcari 1827")])
    csv_path = tmp_path / "wines.csv"
    written = write_csv(csv_path, rows, COLUMNS)
    assert written == len(rows)

    with csv_path.open(encoding="utf-8-sig") as fh:
        back = list(csv.DictReader(fh))
    assert len(back) == len(rows)
    assert set(back[0]) == set(COLUMNS)

    xlsx_path = tmp_path / "wines.xlsx"
    write_xlsx(xlsx_path, rows, [], [])
    from openpyxl import load_workbook

    wb = load_workbook(xlsx_path)
    assert wb.sheetnames == [
        "Wines", "Needs review", "Duplicates", "Unmatched tags", "Errors"
    ]
    assert wb["Wines"].max_row == 3  # header + 2 primary rows


# --------------------------------------------------------------------------
# Regressions
# --------------------------------------------------------------------------


def test_tall_band_is_split_vertically_not_downscaled():
    """A band taller than the limit cannot be fixed by horizontal splits alone."""
    tall = _band(bottles_y0=0.10, bottles_y1=0.68, tags_y0=0.70, tags_y1=0.76,
                 x0=0.05, x1=0.95)
    regions = band_regions(PHOTO_W, PHOTO_H, [tall], MAX_EDGE_HIRES)
    assert len(regions) > 2, "expected a 2-D split"
    for r in regions:
        assert max(r.width, r.height) <= MAX_EDGE_HIRES, r


def test_vertical_split_marks_which_tiles_hold_the_tag_rail():
    tall = _band(bottles_y0=0.10, bottles_y1=0.68, tags_y0=0.70, tags_y1=0.76,
                 x0=0.05, x1=0.95)
    regions = band_regions(PHOTO_W, PHOTO_H, [tall], MAX_EDGE_HIRES)
    flags = {r.contains_tag_rail for r in regions}
    assert flags == {True, False}, "some tiles hold the rail, some do not"
    # The rail sits at the foot of a band, so only the lowest tiles carry it.
    rail_top = min(r.y0 for r in regions if r.contains_tag_rail)
    assert all(r.y0 <= rail_top for r in regions if not r.contains_tag_rail)


def test_short_band_keeps_the_rail_in_every_tile():
    regions = band_regions(PHOTO_W, PHOTO_H, [_band(x0=0.0, x1=1.0)], MAX_EDGE_HIRES)
    assert len(regions) > 1
    assert all(r.contains_tag_rail for r in regions)


def test_grid_tiling_leaves_rail_membership_unknown():
    # Grid mode has no layout information, so it must not assert either way.
    assert all(r.contains_tag_rail is None for r in grid_regions(PHOTO_W, PHOTO_H))


def test_tag_free_tiles_are_told_not_to_price_bottles():
    from wine_ocr.prompts import extract_user_prompt

    without = extract_user_prompt("p.jpg", "t", None, "RON", None, True, False)
    assert "leave price null" in without

    for flag in (True, None):
        text = extract_user_prompt("p.jpg", "t", None, "RON", None, True, flag)
        assert "leave price null" not in text


def test_many_sightings_get_distinct_duplicate_ids():
    """Two overlapping photos x two overlapping tiles = four sightings."""
    rows = deduplicate([_row() for _ in range(4)])
    ids = [r["row_id"] for r in rows]
    assert len(set(ids)) == len(ids) == 4
    assert sum(1 for r in rows if not r["duplicate_of"]) == 1
