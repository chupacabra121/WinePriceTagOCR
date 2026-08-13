"""Turn extraction results into rows, then into CSV / Excel / JSONL."""

from __future__ import annotations

import csv
import json
from dataclasses import asdict, is_dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Iterable, Optional

from .extract import PhotoResult
from .images import PhotoMeta
from .normalize import (
    check_price,
    check_unit_price,
    detect_currency,
    dedupe_key,
    discount_pct,
    normalize_vintage,
    normalize_volume,
    price_per_litre,
    review_reasons,
)
from .stores import StoreResolution

COLUMNS = [
    "row_id", "store", "wine_name", "producer", "vintage", "wine_type", "sweetness",
    "price", "currency", "volume_ml", "price_per_litre", "price_kind",
    "original_price", "discount_pct", "grape_varieties", "region", "country",
    "abv_percent", "promo_text", "unit_price_text",
    "needs_review", "review_reasons", "pairing_confidence", "pairing_note",
    "price_check", "unit_price_check", "name_source",
    "photo", "photo_taken_at", "store_source", "store_read_from_photo",
    "shelf", "price_text", "raw_tag_text", "raw_label_text",
    "gps_lat", "gps_lon", "bottle_bbox", "tag_bbox",
    "photo_sha256", "model", "extracted_at", "duplicate_of",
]

NUMERIC = {"price", "price_per_litre", "original_price", "discount_pct",
           "volume_ml", "abv_percent", "gps_lat", "gps_lon"}

UNMATCHED_COLUMNS = [
    "store", "photo", "shelf", "raw_text", "price", "currency", "reason", "tag_bbox"
]


def _fmt_bbox(bbox: Optional[tuple[float, float, float, float]]) -> str:
    if not bbox:
        return ""
    return ",".join(f"{v:.4f}" for v in bbox)


def rows_from_photo(
    result: PhotoResult,
    meta: PhotoMeta,
    store: StoreResolution,
    model: str,
    root: Path,
) -> tuple[list[dict], list[dict]]:
    """Flatten one photo's regions into wine rows and unmatched-tag rows."""
    extracted_at = datetime.now(timezone.utc).isoformat(timespec="seconds")
    try:
        rel_photo = str(meta.path.relative_to(root))
    except ValueError:
        rel_photo = str(meta.path)

    layout_currency = result.layout.currency_guess if result.layout else None
    rows: list[dict] = []
    unmatched: list[dict] = []

    for region_result in result.regions:
        extraction = region_result.extraction
        if extraction is None:
            continue
        region = region_result.region

        for wine in extraction.wines:
            volume = normalize_volume(
                wine.volume_ml, wine.raw_tag_text, wine.raw_label_text, wine.wine_name
            )
            currency = wine.currency or layout_currency or detect_currency(
                wine.price_text, wine.raw_tag_text
            )
            vintage = normalize_vintage(
                wine.vintage, wine.raw_label_text, wine.raw_tag_text
            )
            row = {
                "row_id": "",  # assigned after dedup
                "store": store.store,
                "wine_name": wine.wine_name.strip(),
                "producer": wine.producer,
                "vintage": vintage,
                "wine_type": wine.wine_type,
                "sweetness": wine.sweetness,
                "price": wine.price,
                "currency": currency,
                "volume_ml": volume,
                "price_per_litre": price_per_litre(wine.price, volume),
                "price_kind": wine.price_kind,
                "original_price": wine.original_price,
                "discount_pct": discount_pct(wine.price, wine.original_price),
                "grape_varieties": "; ".join(wine.grape_varieties),
                "region": wine.region,
                "country": wine.country,
                "abv_percent": wine.abv_percent,
                "promo_text": wine.promo_text,
                "unit_price_text": wine.unit_price_text,
                "pairing_confidence": wine.pairing_confidence,
                "pairing_note": wine.pairing_note,
                "price_check": check_price(wine.price, wine.price_text),
                "unit_price_check": check_unit_price(
                    price_per_litre(wine.price, volume), wine.unit_price_text
                ),
                "name_source": wine.name_source,
                "photo": rel_photo,
                "photo_taken_at": meta.taken_at,
                "store_source": store.source,
                "store_read_from_photo": store.read_from_photo,
                "shelf": region.label,
                "price_text": wine.price_text,
                "raw_tag_text": wine.raw_tag_text,
                "raw_label_text": wine.raw_label_text,
                "gps_lat": meta.gps_lat,
                "gps_lon": meta.gps_lon,
                "bottle_bbox": _fmt_bbox(
                    region.to_full(
                        wine.bottle_bbox.x0, wine.bottle_bbox.y0,
                        wine.bottle_bbox.x1, wine.bottle_bbox.y1,
                        meta.width, meta.height,
                    ) if wine.bottle_bbox else None
                ),
                "tag_bbox": _fmt_bbox(
                    region.to_full(
                        wine.tag_bbox.x0, wine.tag_bbox.y0,
                        wine.tag_bbox.x1, wine.tag_bbox.y1,
                        meta.width, meta.height,
                    ) if wine.tag_bbox else None
                ),
                "photo_sha256": meta.sha256,
                "model": model,
                "extracted_at": extracted_at,
                "duplicate_of": "",
            }
            reasons = review_reasons(row)
            row["needs_review"] = "yes" if reasons else ""
            row["review_reasons"] = "; ".join(reasons)
            rows.append(row)

        for tag in extraction.unreadable_tags:
            unmatched.append(
                {
                    "store": store.store,
                    "photo": rel_photo,
                    "shelf": region.label,
                    "raw_text": tag.raw_text,
                    "price": tag.price,
                    "currency": tag.currency or layout_currency,
                    "reason": tag.reason,
                    "tag_bbox": _fmt_bbox(
                        region.to_full(
                            tag.tag_bbox.x0, tag.tag_bbox.y0,
                            tag.tag_bbox.x1, tag.tag_bbox.y1,
                            meta.width, meta.height,
                        ) if tag.tag_bbox else None
                    ),
                }
            )

    return rows, unmatched


def deduplicate(rows: list[dict]) -> list[dict]:
    """Collapse the same product seen more than once.

    Two things cause duplicates: overlapping tiles within one photo, and
    overlapping photos of the same shelf. Both are handled the same way — keep
    the most informative row and point the rest at it via ``duplicate_of``
    rather than deleting them, so nothing silently disappears.
    """

    def informativeness(row: dict) -> tuple:
        return (
            row.get("price") is not None,
            {"high": 2, "medium": 1, "low": 0}.get(row.get("pairing_confidence"), 0),
            row.get("volume_ml") is not None,
            row.get("vintage") is not None,
            len(row.get("raw_tag_text") or ""),
            len(row.get("wine_name") or ""),
        )

    groups: dict[str, list[dict]] = {}
    for row in rows:
        key = dedupe_key(
            row["store"], row["wine_name"], row.get("vintage"), row.get("volume_ml")
        )
        groups.setdefault(key, []).append(row)

    kept: list[dict] = []
    for index, (_, group) in enumerate(sorted(groups.items()), start=1):
        group.sort(key=informativeness, reverse=True)
        primary = group[0]
        primary["row_id"] = f"W{index:05d}"
        kept.append(primary)

        prices = {r["price"] for r in group if r.get("price") is not None}
        if len(prices) > 1:
            extra = f"price disagrees across sightings ({sorted(prices)})"
            primary["review_reasons"] = "; ".join(
                filter(None, [primary["review_reasons"], extra])
            )
            primary["needs_review"] = "yes"

        # Suffix every duplicate distinctly: one wine is routinely seen 3-4
        # times (two overlapping photos x two overlapping tiles), and a shared
        # id would collide.
        for n, dup in enumerate(group[1:], start=1):
            dup["row_id"] = f"{primary['row_id']}d{n}"
            dup["duplicate_of"] = primary["row_id"]
            kept.append(dup)

    return kept


# --------------------------------------------------------------------------
# Writers
# --------------------------------------------------------------------------


def write_csv(path: Path, rows: Iterable[dict], columns: list[str]) -> int:
    path.parent.mkdir(parents=True, exist_ok=True)
    count = 0
    with path.open("w", encoding="utf-8-sig", newline="") as fh:
        writer = csv.DictWriter(fh, fieldnames=columns, extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            writer.writerow({c: row.get(c, "") for c in columns})
            count += 1
    return count


def write_jsonl(path: Path, records: Iterable[Any]) -> int:
    path.parent.mkdir(parents=True, exist_ok=True)
    count = 0
    with path.open("w", encoding="utf-8") as fh:
        for record in records:
            payload = asdict(record) if is_dataclass(record) else record
            fh.write(json.dumps(payload, ensure_ascii=False, default=str) + "\n")
            count += 1
    return count


def write_xlsx(
    path: Path,
    rows: list[dict],
    unmatched: list[dict],
    errors: list[dict],
) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="6B2737")  # wine red
    review_fill = PatternFill("solid", fgColor="FFF3CD")

    def sheet(title: str, columns: list[str], data: list[dict], highlight: bool = False):
        ws = wb.create_sheet(title)
        ws.append(columns)
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(vertical="center")
        for row in data:
            ws.append([row.get(c, "") for c in columns])

        if highlight and data:
            flag_col = columns.index("needs_review") + 1
            for r in range(2, len(data) + 2):
                if ws.cell(row=r, column=flag_col).value == "yes":
                    for c in range(1, len(columns) + 1):
                        ws.cell(row=r, column=c).fill = review_fill

        for i, name in enumerate(columns, start=1):
            letter = get_column_letter(i)
            longest = max(
                [len(name)] + [len(str(row.get(name, "") or "")) for row in data[:400]]
            )
            ws.column_dimensions[letter].width = min(max(10, longest + 2), 46)
            if name in NUMERIC:
                fmt = "#,##0.00" if name != "volume_ml" else "#,##0"
                for r in range(2, len(data) + 2):
                    ws.cell(row=r, column=i).number_format = fmt

        ws.freeze_panes = "A2"
        if data:
            ws.auto_filter.ref = (
                f"A1:{get_column_letter(len(columns))}{len(data) + 1}"
            )
        return ws

    primary = [r for r in rows if not r.get("duplicate_of")]
    sheet("Wines", COLUMNS, primary, highlight=True)
    sheet("Needs review", COLUMNS, [r for r in primary if r.get("needs_review") == "yes"])
    sheet("Duplicates", COLUMNS, [r for r in rows if r.get("duplicate_of")])
    sheet("Unmatched tags", UNMATCHED_COLUMNS, unmatched)
    sheet("Errors", ["photo", "stage", "detail"], errors)

    wb.remove(wb["Sheet"])
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
